import re
from datetime import date as date_type, timedelta
import pandas as pd

DATE_FROM_FILENAME = re.compile(r'(\d{4})-(\d{2})-(\d{2})')
import numpy as np
import random
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def fix_date(col):
    def conv(v):
        if pd.isna(v):
            return pd.NaT
        if isinstance(v, (int, float)):
            return pd.Timestamp("1899-12-30") + pd.Timedelta(days=float(v))
        return pd.Timestamp(v)
    return col.apply(conv)


def ordinal(n):
    """Return ordinal string for an integer, e.g. 1 -> '1st', 11 -> '11th', 22 -> '22nd'."""
    if 11 <= (n % 100) <= 13:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"

def format_date_range(dates):
    """Return a date range string in M-D format like '6-1 thru 6-16' (or '6-1' for a single day)."""
    sorted_dates = sorted(dates)
    def fmt(d):
        return f"{d.month}-{d.day}"
    if len(sorted_dates) == 1:
        return fmt(sorted_dates[0])
    return f"{fmt(sorted_dates[0])} thru {fmt(sorted_dates[-1])}"


VENDOR_REPLACEMENTS = [
    ("AXS", "Veritix"), ("AXS.com", "Veritix"),
    ("Ticketmaster.com", "Ticketmaster"), ("Toyota Center- Houston", "Veritix"),
    ("Ak-Chin Pavilion", "Live Nation Ak-Chin Pavilion"),
    ("BB&T Pavilion", "Live Nation BB&T Pavilion"),
    ("Dos Equis Pavilion", "Live Nation Dos Equis Pavilion"),
    ("Jiffy Lube Live", "Live Nation Jiffy Lube Live"),
    ("Jacobs Pavilion", "Live Nation Jacobs Pavilion"),
    ("MIDFLORIDA Credit Union Amp", "Live Nation MidFlorida Credit Union Amp"),
    ("Shoreline Amphitheatre", "Live Nation Shoreline Amphitheatre"),
    ("Ticketmaster Phones", "Ticketmaster"),
    ("Shubert Organization Telecharge", "Telecharge"),
    ("Veritix.com", "Veritix"),
    ("Huntington Bank Pavilion", "Live Nation Huntington Bank"),
    ("Darien Lake Amphitheatre", "Live Nation Darien Lake Amphitheatre"),
    ("Leader Bank Pavillion", "Live Nation Leader Bank Pavilion"),
    ("Coastal Credit Union Music Park at Walnut Creek", "Live Nation MidFlorida Credit Union Amp"),
    ("TD Pavilion at the Mann", "Live Nation TD Pavilion at the Mann"),
    ("Live Nation", "Ticketmaster"),
    ("Xfinity Theatre", "Live Nation Xfinity Theatre CT"),
    ("Live Nation Xfinity Boston", "Live Nation Xfinity Center Boston"),
    ("Live Nation PNC Bank", "Live Nation PNC Bank Charlotte"),
    ("LN Ruoff Home Mortgage Music Center", "Live Nation Ruoff"),
    ("Alpine Valley", "Live Nation Alpine Valley"),
    ("The Pavilion at Toyota Music Factory", "Live Nation Pavilion at Toyota Music Factory"),
    ("Live Nation Pnc Charlotte", "Live Nation PNC Music Pavilion"),
    ("The Pavilion at Star Lake", "Live Nation Pavilion At Star Lake"),
    ("Bank of New Hampshire Pavilion", "Live Nation Bank of New Hampshire"),
    ("North Island Credit Union", "Live Nation North Island Credit Union"),
    ("Northwell Health at Jones Beach Theater", "Live Nation Jones Beach"),
    ("PNC Music Pavilion", "Live Nation PNC Music Pavilion"),
    ("Ruoff Music Center", "Live Nation Ruoff Music Center"),
    ("Waterfront Music Pavilion", "Live Nation BB&T Pavilion"),
    ("Concord Pavilion", "Live Nation Concord Pavilion"),
    ("Bethel Woods", "Live Nation Bethel Woods"),
    ("Ford Idaho Center Ampitheater", "Live Nation Ford Idaho Amp"),
    ("The Met Philadelphia", "Live Nation The Met Philadelphia"),
    ("The Masonic", "Live Nation Masonic"),
    ("FPL Solar Ampitheater at Bayfront Park", "Live Nation FPL Solar Amp"),
    ("Xfinity Center", "Live Nation Xfinity Center Boston"),
    ("The Cynthia Woods Mitchell Pavilion", "Live Nation Cynthia Woods Mitchell Pavilion"),
    ("Toyota Amphitheatre", "Live Nation Toyota Amp"),
    ("Cellairis Amphitheatre at Lakewood", "Live Nation Cellairis"),
    ("Oak Mountian Amphitheatre", "Live Nation Oak Mountain"),
    ("Chicago Fire", "Chicago Fire FC"), ("Dallas FC", "FC Dallas"),
    ("Houston Dynamo", "Houston Dynamo FC"), ("LAFC", "Los Angeles FC"),
    ("LA Galaxy", "Los Angeles Galaxy"), ("Minnesota United", "Minnesota United FC"),
    ("New York FC", "New York City FC"), ("Seattle Sounders", "Seattle Sounders FC"),
    ("St. Louis City", "St. Louis City SC"), ("Vancouver Whitecaps", "Vancouver Whitecaps FC"),
    ("Portland Trailblazers", "Portland Trail Blazers"),
    ("Concert Extras", "Live Nation Extras"),
    ("PHILADELPHIA 76ERS", "Philadelphia 76ers"),
    ("SAN FRANCISCO 49RS", "San Francisco 49ers"),
    ("Concert Partials", "Concert Seasons"), ("Legacy StubHub", "StubHub"),
]

BROADWAY_VENUES = {
    "Brooks Atkinson Theatre": "Box Office - Brooks Atkinson Theatre",
    "Hudson Broadway Theatre": "Box Office - Hudson Broadway Theatre",
    "Minskoff Theatre": "Box Office - Minskoff Theatre",
    "Madison Square Garden": "Box Office - MSG Advance",
    "Neil Simon Theatre": "Box Office - Neil Simon Theatre",
    "Richard Rodgers Theatre": "Box Office - Richard Rodgers Theatre",
    "Winter Garden Theater New York": "Box Office - Winter Garden Theatre",
    "Walter Kerr Theatre": "Box Office - Walter Kerr Theatre",
    "August Wilson Theatre": "Box Office - August Wilson Theatre",
    "Music Box Theatre New York": "Box Office - Music Box Theatre",
    "Imperial Theatre New York": "Box Office - Imperial Theatre",
    "Lyceum Theatre New York": "Box Office - Lyceum Theatre",
    "Booth Theatre": "Box Office - Booth Theatre",
    "Longacre Theatre": "Box Office - Longacre Theatre",
    "Majestic Theatre New York": "Box Office - Majestic Theatre",
    "Broadhurst Theatre": "Box Office - Broadhurst Theatre",
    "Stephen Sondheim Theatre": "Box Office - Stephen Sondheim Theatre",
    "Lena Horne Theatre": "Box Office - Lena Horne Theatre",
    "Lyric Theatre - NY": "Box Office - Lyric Theatre",
    "Winter Garden Theatre (Toronto)": "Box Office - Winter Garden Theatre (Toronto)",
    "Marquis Theatre New York": "Box Office - Marquis Theatre",
    "Lunt Fontanne Theatre": "Box Office - Lunt Fontanne Theatre",
    "John Golden Theatre": "Box Office - John Golden Theatre",
    "Circle In The Square": "Box Office - Circle In The Square",
}

CONCERT_SEASONS_MAP = {
    "Ak-Chin Pavilion": "Live Nation Ak-Chin Pavilion",
    "Alpine Valley Music Theatre": "Live Nation Alpine Valley",
    "Bank of New Hampshire Pavilion": "Live Nation Bank of New Hampshire",
    "Darling's Waterfront Pavilion": "Live Nation Waterfront",
    "Darlings Waterfront Pavilion": "Live Nation Waterfront",
    "Bethel Woods Center For The Arts": "Live Nation Bethel Woods",
    "Blossom Music Center": "Live Nation Blossom MC",
    "Lakewood Amphitheatre": "Live Nation Lakewood Amphitheatre",
    "Coastal Credit Union Music Park at Walnut Creek": "Live Nation Coastal Credit Union",
    "Concord Pavilion": "Live Nation Concord Pavilion",
    "Cynthia Woods Mitchell Pavilion": "Live Nation Cynthia Woods Mitchell Pavilion",
    "Darien Lake Amphitheater": "Live Nation Darien Lake Amphitheatre",
    "Dos Equis Pavilion": "Live Nation Dos Equis Pavilion",
    "FivePoint Amphitheatre": "Live Nation FivePoint",
    "Ford Idaho Center": "Live Nation Ford Idaho Amp",
    "Bayfront Park Amphitheatre": "Live Nation FPL Solar Amp",
    "Glen Helen Amphitheater": "Live Nation Glen Helen",
    "Gorge Amphitheatre": "Live Nation Gorge",
    "Hershey Theatre": "Live Nation Hershey",
    "Hollywood Casino Amphitheatre - Tinley Park": "Live Nation Hollywood Casino - Tinley Park",
    "Huntington Bank Pavilion at Northerly Island": "Live Nation Huntington Bank",
    "Isleta Amphitheater": "Live Nation Isleta Amphitheatre",
    "Jacobs Pavilion at Nautica": "Live Nation Jacobs Pavilion",
    "Jiffy Lube Live": "Live Nation Jiffy Lube Live",
    "Northwell Health at Jones Beach Theater": "Live Nation Jones Beach",
    "KeyBank Center": "Live Nation KeyBank",
    "Leader Bank Pavilion": "Live Nation Leader Bank Pavilion",
    "The Masonic": "Live Nation Masonic",
    "MGM Music Hall at Fenway": "Live Nation MGM Music Hall",
    "MidFlorida Credit Union Amphitheatre": "Live Nation MidFlorida Credit Union Amp",
    "North Island Credit Union Amphitheatre": "Live Nation North Island Credit Union",
    "Oak Mountain Amphitheatre": "Live Nation Oak Mountain",
    "The Pavilion at Star Lake": "Live Nation Pavilion At Star Lake",
    "Pavilion at the Toyota Music Factory": "Live Nation Pavilion At Toyota Music Factory",
    "PNC Bank Arts Center": "Live Nation PNC Bank Charlotte",
    "PNC Music Pavilion": "Live Nation PNC Music Pavilion",
    "Ruoff Music Center": "Live Nation Ruoff",
    "Shoreline Amphitheatre": "Live Nation Shoreline Amphitheatre",
    "TD Pavilion at the Mann": "Live Nation TD Pavilion at the Mann",
    "The Met Philadelphia": "Live Nation The Met Philadelphia",
    "Toyota Amphitheatre": "Live Nation Toyota Amp",
    "White River Amphitheatre": "Live Nation White River",
    "The Wiltern": "Live Nation Wiltern",
    "Xfinity Center": "Live Nation Xfinity Center Boston",
    "Xfinity Theatre": "Live Nation Xfinity Theatre CT",
    "Freedom Mortgage Pavilion": "Live Nation Freedom Mortgage Music Pavilion",
    "Bayfront Park-Miami": "Live Nation FPL Solar Amp",
    "Idaho Center Amphitheater": "Live Nation Ford Idaho Amp",
    "Coca-Cola Roxy": "Live Nation Coca-Cola Roxy Theatre",
    "Farm Bureau Insurance Lawn at White River State Park": "Live Nation TCU Amp",
    "Charlotte Metro Credit Union Amphitheatre": "Live Nation Charlotte Metro Credit Union",
    "713 Music Hall": "Live Nation 713 Music Hall",
    "TCU Amphitheater at White River State Park": "Live Nation TCU Amp",
    "USANA Amphitheatre": "Live Nation USANA Amp",
    "Hollywood Casino Amphitheater - St. Louis": "Live Nation Hollywood Casino - St. Louis",
    "Hollywood Casino Amphitheatre St Louis": "Live Nation Hollywood Casino - St. Louis",
    "iTHINK Financial Amphitheatre": "Live Nation iTHINK Financial Amp",
    "Ameris Bank Amphitheatre": "Live Nation Ameris Bank Amp",
    "Ascend Amphitheater": "Live Nation Ascend Amp",
    "FirstBank Amphitheater": "Live Nation FirstBank Amp",
    "Hersheypark Stadium": "Live Nation Hersheypark Stadium",
    "Veterans United Home Loans Amphitheater": "Live Nation Veterans United Amp",
    "Starlight Theatre": "Live Nation Starlight Theatre",
    "Riverbend Music Center": "Live Nation Riverbend Music Center",
    "The Terminal - Houston": "Live Nation 713 Music Hall",
    "Veterans United Home Loans Amphitheater at Virginia Beach": "Live Nation Veterans United Amp",
    "Saint Louis Music Park": "Live Nation Saint Louis Music Park",
    "Old National Centre": "Live Nation Old National Centre",
    "Skyla Credit Union": "Live Nation Skyla Credit Union Amp",
    "Skyla Credit Union Amphitheatre": "Live Nation Skyla Credit Union Amp",
    "Pine Knob Music Center": "Live Nation Pine Knob Music Center",
    "St. Joseph's Health Amphitheater at Lakeview": "Live Nation St. Joseph's Health Amp",
    "Red Hat Amphitheater": "Live Nation Red Hat Amphitheater",
    "Talking Stick Resort Amphitheatre": "Live Nation Talking Stick Resort Amp",
    "The Fillmore Detroit": "Live Nation The Fillmore Detroit",
    "CFG Bank Arena": "Live Nation CFG Bank Arena",
    "Aragon Ballroom": "Live Nation Aragon Ballroom",
    "Saratoga Performing Arts Center": "Live Nation Saratoga Springs PAC",
    "The Fillmore - Charlotte": "Live Nation The Fillmore Charlotte",
    "Fillmore Auditorium-CO": "Live Nation The Fillmore Denver",
    "The Fillmore - Philadelphia": "Live Nation The Fillmore Philly",
    "The Fillmore-Silver Spring": "Live Nation The Fillmore Silver Spring",
    "SAP Center at San Jose": "Live Nation SAP Center at San Jose",
    "Arizona Federal Theatre": "Live Nation Arizona Federal Theatre",
    "Flagstar at Westbury Music Fair": "Live Nation Flagstar At Westbury Music Fair",
    "Uptown Minneapolis": "Live Nation Uptown Minneapolis",
    "The Pavilion At Toyota Music Factory": "Live Nation Pavilion At Toyota Music Factory",
    "Forest Hills Stadium": "Forest Hills Stadium",
    "NRG Stadium": "Houston Rodeo",
    "Hayden Homes Amphitheater": "Live Nation Hayden Homes Amphitheater",
    "The Dome at Oakdale Theatre": "Live Nation Toyota Oakdale Theatre",
    "The Dome at Toyota Oakdale Theatre": "Live Nation Toyota Oakdale Theatre",
    "Broadview Stage at SPAC": "Live Nation Broadview Stage at SPAC",
    "BECU Live Outdoor Venue": "Live Nation BECU",
    "BankNH Pavilion": "Live Nation Bank of New Hampshire",
    "Everwise Amphitheater at White River State Park": "Live Nation White River",
    "The Cynthia Woods Mitchell Pavilion presented by Huntsman": "Live Nation Cynthia Woods Mitchell Pavilion",
    "Harbor Yard Amphitheater": "Live Nation Harbor Yard Amp",
    "Toyota Pavilion at Concord": "Live Nation Concord Pavilion",
    "Utah First Credit Union Amphitheatre (formerly USANA Amp)": "Live Nation Usana Amp",
    "Toyota Oakdale Theatre": "Live Nation Toyota Oakdale Theatre",
    "Byline Bank Aragon Ballroom": "Live Nation Aragon Ballroom",
    "Skyline Stage at the Mann": "Live Nation Skyline Stage At The Mann",
    "Santa Barbara Bowl": "Live Nation Santa Barbara Bowl",
    "20 Monroe Live": "Live Nation GLC Live at 20 Monroe",
    "MIDFLORIDA Credit Union Amphitheatre at the FL State Fairgrounds": "Live Nation MidFlorida Credit Union Amp",
    "Credit Union 1 Amphitheatre": "Live Nation Credit Union 1 Amphitheatre",
    "Daily's Place": "Live Nation Dailys Place",
    "Greek Theatre Los Angeles": "Live Nation Greek Theatre Los Angeles",
    "Koka Booth Field 2": "Live Nation Koka Booth Amphitheatre",
    "Michigan Lottery Amphitheatre at Freedom Hill": "Live Nation Michigan Lottery Amphitheatre",
    "Mountain Winery": "Live Nation Mountain Winery",
    "Skyla Credit Union Amphitheatre at AvidXchange Music Factory": "Live Nation Skyla Credit Union Amp",
    "Vibrant Music Hall": "Live Nation Vibrant Music Hall",
    "Vina Robles Amphitheatre": "Live Nation Vina Robles Amphitheatre",
    "Whitewater Amphitheater": "Live Nation Whitewater Amphitheater",
    "Old National Centre Complex": "Live Nation Old National Centre",
    "Fiddlers Green Amphitheatre": "Live Nation Fiddlers Green Amphitheatre",
    "Pine Knob Music Theatre": "Live Nation Pine Knob Music Center",
    "Hollywood Palladium": "Live Nation Hollywood Palladium",
    "Hartford HealthCare Amphitheater": "Live Nation Hartford Healthcare Amphitheater (Harbor Yard)",
    "Mystic Lake Amphitheater": "Live Nation Mystic Lake Amphitheater",
    "Fillmore Minneapolis": "Live Nation Fillmore Minneapolis",
    "Uptown Theater - Minneapolis": "Live Nation Uptown Minneapolis",
    "Old National Centre.": "Live Nation Old National Centre",
    "Sandy Amphitheater": "Live Nation Sandy Amphitheater",
}

BROADWAY_SEASONS_MAP = {
    "Boston Opera House": "Broadway Boston",
    "Colonial Theatre Boston": "Broadway Boston",
    "Fox Theatre - Atlanta": "Broadway Atlanta",
    "Hippodrome Theatre": "Broadway Baltimore",
    "BJCC Concert Hall": "Broadway Birmingham",
    "Shea's Buffalo Theatre": "Broadway Buffalo",
    "Procter and Gamble Hall at Aronoff Center for the Arts": "Broadway Cincinnati",
    "AT&T Performing Arts Center - Winspear Opera House": "Broadway Dallas",
    "Music Hall at Fair Park": "Broadway Dallas",
    "Durham Performing Arts Center": "Broadway Durham",
    "Broward Center Amaturo": "Broadway Ft Lauderdale",
    "Devos Performance Hall": "Broadway Grand Rapids",
    "Hollywood Pantages Theatre": "Broadway Hollywood",
    "Music Hall - Kansas City": "Broadway Kansas City",
    "Muriel Kauffman Theatre at Kauffman Center for the Performing Arts": "Broadway Kansas City",
    "Saenger Theatre-New Orleans": "Broadway New Orleans",
    "Sarofim Hall at The Hobby Center": "Broadway Houston",
    "Uihlein Hall at Marcus Center for the Performing Arts": "Broadway Milwaukee",
    "Orpheum Theatre Minneapolis": "Broadway Minneapolis",
    "Clowes Memorial Hall": "Broadway Indianapolis",
    "Old National Centre": "Broadway Indianapolis",
    "Paramount Theatre": "Broadway Seattle",
    "San Diego Civic Theatre": "Broadway San Diego",
    "San Jose Center for the Performing Arts": "Broadway San Jose",
}

MLB_TEAMS = {
    "Anaheim Ducks",
    "Arizona Cardinals",
    "Arizona Diamondbacks",
    "Athletics",
    "Atlanta Braves",
    "Atlanta Falcons",
    "Atlanta Hawks",
    "Atlanta United FC",
    "Austin FC",
    "Baltimore Orioles",
    "Baltimore Ravens",
    "Boston Bruins",
    "Boston Celtics",
    "Boston Red Sox",
    "Brooklyn Nets",
    "Buffalo Bills",
    "Buffalo Sabres",
    "CF Montreal",
    "Calgary Flames",
    "Carolina Hurricanes",
    "Carolina Panthers",
    "Charlotte FC",
    "Charlotte Hornets",
    "Chicago Bears",
    "Chicago Blackhawks",
    "Chicago Bulls",
    "Chicago Cubs",
    "Chicago Fire",
    "Chicago White Sox",
    "Cincinnati Bengals",
    "Cincinnati Reds",
    "Cleveland Browns",
    "Cleveland Cavaliers",
    "Cleveland Guardians",
    "Colorado Avalanche",
    "Colorado Rapids",
    "Colorado Rockies",
    "Columbus Blue Jackets",
    "Columbus Crew",
    "DC United",
    "Dallas Cowboys",
    "Dallas Mavericks",
    "Dallas Stars",
    "Denver Broncos",
    "Denver Nuggets",
    "Detroit Lions",
    "Detroit Pistons",
    "Detroit Red Wings",
    "Detroit Tigers",
    "Edmonton Oilers",
    "FC Cincinnati",
    "FC Dallas",
    "Florida Panthers",
    "Golden State Warriors",
    "Green Bay Packers",
    "Houston Astros",
    "Houston Dynamo",
    "Houston Rockets",
    "Houston Texans",
    "Indiana Pacers",
    "Indianapolis Colts",
    "Inter Miami CF",
    "Jacksonville Jaguars",
    "Kansas City Chiefs",
    "Kansas City Royals",
    "LA Clippers",
    "Las Vegas Raiders",
    "Los Angeles Angels",
    "Los Angeles Chargers",
    "Los Angeles Dodgers",
    "Los Angeles FC",
    "Los Angeles Galaxy",
    "Los Angeles Kings",
    "Los Angeles Lakers",
    "Los Angeles Rams",
    "Memphis Grizzlies",
    "Miami Dolphins",
    "Miami Heat",
    "Miami Marlins",
    "Milwaukee Brewers",
    "Milwaukee Bucks",
    "Minnesota Timberwolves",
    "Minnesota Twins",
    "Minnesota United FC",
    "Minnesota Vikings",
    "Minnesota Wild",
    "Montreal Canadiens",
    "Nashville Predators",
    "Nashville SC",
    "New England Patriots",
    "New England Revolution",
    "New Jersey Devils",
    "New Orleans Pelicans",
    "New Orleans Saints",
    "New York City FC",
    "New York Giants",
    "New York Islanders",
    "New York Jets",
    "New York Knicks",
    "New York Mets",
    "New York Rangers",
    "New York Red Bulls",
    "New York Yankees",
    "Oklahoma City Thunder",
    "Orlando City SC",
    "Orlando Magic",
    "Ottawa Senators",
    "Philadelphia 76ers",
    "Philadelphia Eagles",
    "Philadelphia Flyers",
    "Philadelphia Phillies",
    "Philadelphia Union",
    "Phoenix Suns",
    "Pittsburgh Penguins",
    "Pittsburgh Pirates",
    "Pittsburgh Steelers",
    "Portland Timbers",
    "Portland Trail Blazers",
    "Real Salt Lake",
    "Sacramento Kings",
    "Saint Louis City SC",
    "San Antonio Spurs",
    "San Diego FC",
    "San Diego Padres",
    "San Francisco 49ers",
    "San Francisco Giants",
    "San Jose Earthquakes",
    "San Jose Sharks",
    "Seattle Kraken",
    "Seattle Mariners",
    "Seattle Seahawks",
    "Seattle Sounders",
    "Sporting Kansas City",
    "St. Louis Blues",
    "St. Louis Cardinals",
    "Tampa Bay Buccaneers",
    "Tampa Bay Lightning",
    "Tampa Bay Rays",
    "Tennessee Titans",
    "Texas Rangers",
    "Toronto Blue Jays",
    "Toronto FC",
    "Toronto Maple Leafs",
    "Toronto Raptors",
    "Utah Jazz",
    "Utah Mammoth",
    "Vancouver Canucks",
    "Vancouver Whitecaps FC",
    "Vegas Golden Knights",
    "Washington Capitals",
    "Washington Commanders",
    "Washington Nationals",
    "Washington Wizards",
    "Winnipeg Jets"
}
TEAM_LEAGUE = {
    'Anaheim Ducks': 'NHL',
    'Arizona Cardinals': 'NFL',
    'Arizona Diamondbacks': 'MLB',
    'Athletics': 'MLB',
    'Atlanta Braves': 'MLB',
    'Atlanta Dream': 'WNBA',
    'Atlanta Falcons': 'NFL',
    'Atlanta Hawks': 'NBA',
    'Atlanta United FC': 'MLS',
    'Austin FC': 'MLS',
    'Baltimore Orioles': 'MLB',
    'Baltimore Ravens': 'NFL',
    'Boston Bruins': 'NHL',
    'Boston Celtics': 'NBA',
    'Boston Red Sox': 'MLB',
    'Brooklyn Nets': 'NBA',
    'Buffalo Bills': 'NFL',
    'Buffalo Sabres': 'NHL',
    'CF Montreal': 'MLS',
    'Calgary Flames': 'NHL',
    'Calgary Stampeders': 'CFL',
    'Carolina Hurricanes': 'NHL',
    'Carolina Panthers': 'NFL',
    'Charlotte FC': 'MLS',
    'Charlotte Hornets': 'NBA',
    'Chicago Bears': 'NFL',
    'Chicago Blackhawks': 'NHL',
    'Chicago Bulls': 'NBA',
    'Chicago Cubs': 'MLB',
    'Chicago Fire': 'MLS',
    'Chicago Sky': 'WNBA',
    'Chicago White Sox': 'MLB',
    'Cincinnati Bengals': 'NFL',
    'Cincinnati Reds': 'MLB',
    'Cleveland Browns': 'NFL',
    'Cleveland Cavaliers': 'NBA',
    'Cleveland Guardians': 'MLB',
    'Colorado Avalanche': 'NHL',
    'Colorado Rapids': 'MLS',
    'Colorado Rockies': 'MLB',
    'Columbus Blue Jackets': 'NHL',
    'Columbus Crew': 'MLS',
    'Connecticut Sun': 'WNBA',
    'DC United': 'MLS',
    'Dallas Cowboys': 'NFL',
    'Dallas Mavericks': 'NBA',
    'Dallas Stars': 'NHL',
    'Dallas Wings': 'WNBA',
    'Denver Broncos': 'NFL',
    'Denver Nuggets': 'NBA',
    'Detroit Lions': 'NFL',
    'Detroit Pistons': 'NBA',
    'Detroit Red Wings': 'NHL',
    'Detroit Tigers': 'MLB',
    'Edmonton Oilers': 'NHL',
    'FC Cincinnati': 'MLS',
    'FC Dallas': 'MLS',
    'Florida Panthers': 'NHL',
    'Golden State Valkyries': 'WNBA',
    'Golden State Warriors': 'NBA',
    'Green Bay Packers': 'NFL',
    'Houston Astros': 'MLB',
    'Houston Dynamo': 'MLS',
    'Houston Rockets': 'NBA',
    'Houston Texans': 'NFL',
    'Indiana Fever': 'WNBA',
    'Indiana Pacers': 'NBA',
    'Indianapolis Colts': 'NFL',
    'Inter Miami CF': 'MLS',
    'Jacksonville Jaguars': 'NFL',
    'Kansas City Chiefs': 'NFL',
    'Kansas City Royals': 'MLB',
    'LA Clippers': 'NBA',
    'Las Vegas Aces': 'WNBA',
    'Las Vegas Raiders': 'NFL',
    'Los Angeles Angels': 'MLB',
    'Los Angeles Chargers': 'NFL',
    'Los Angeles Dodgers': 'MLB',
    'Los Angeles FC': 'MLS',
    'Los Angeles Galaxy': 'MLS',
    'Los Angeles Kings': 'NHL',
    'Los Angeles Lakers': 'NBA',
    'Los Angeles Rams': 'NFL',
    'Los Angeles Sparks': 'WNBA',
    'Memphis Grizzlies': 'NBA',
    'Miami Dolphins': 'NFL',
    'Miami Heat': 'NBA',
    'Miami Marlins': 'MLB',
    'Milwaukee Brewers': 'MLB',
    'Milwaukee Bucks': 'NBA',
    'Minnesota Lynx': 'WNBA',
    'Minnesota Timberwolves': 'NBA',
    'Minnesota Twins': 'MLB',
    'Minnesota United FC': 'MLS',
    'Minnesota Vikings': 'NFL',
    'Minnesota Wild': 'NHL',
    'Montreal Canadiens': 'NHL',
    'Nashville Predators': 'NHL',
    'Nashville SC': 'MLS',
    'New England Patriots': 'NFL',
    'New England Revolution': 'MLS',
    'New Jersey Devils': 'NHL',
    'New Orleans Pelicans': 'NBA',
    'New Orleans Saints': 'NFL',
    'New York City FC': 'MLS',
    'New York Giants': 'NFL',
    'New York Islanders': 'NHL',
    'New York Jets': 'NFL',
    'New York Knicks': 'NBA',
    'New York Liberty': 'WNBA',
    'New York Mets': 'MLB',
    'New York Rangers': 'NHL',
    'New York Red Bulls': 'MLS',
    'New York Yankees': 'MLB',
    'Oklahoma City Thunder': 'NBA',
    'Orlando City SC': 'MLS',
    'Orlando Magic': 'NBA',
    'Ottawa Senators': 'NHL',
    'Philadelphia 76ers': 'NBA',
    'Philadelphia Eagles': 'NFL',
    'Philadelphia Flyers': 'NHL',
    'Philadelphia Phillies': 'MLB',
    'Philadelphia Union': 'MLS',
    'Phoenix Mercury': 'WNBA',
    'Phoenix Suns': 'NBA',
    'Pittsburgh Penguins': 'NHL',
    'Pittsburgh Pirates': 'MLB',
    'Pittsburgh Steelers': 'NFL',
    'Portland Timbers': 'MLS',
    'Portland Trail Blazers': 'NBA',
    'Real Salt Lake': 'MLS',
    'Sacramento Kings': 'NBA',
    'Saint Louis City SC': 'MLS',
    'San Antonio Spurs': 'NBA',
    'San Diego FC': 'MLS',
    'San Diego Padres': 'MLB',
    'San Francisco 49ers': 'NFL',
    'San Francisco Giants': 'MLB',
    'San Jose Earthquakes': 'MLS',
    'San Jose Sharks': 'NHL',
    'Seattle Kraken': 'NHL',
    'Seattle Mariners': 'MLB',
    'Seattle Seahawks': 'NFL',
    'Seattle Sounders': 'MLS',
    'Seattle Storm': 'WNBA',
    'Sporting Kansas City': 'MLS',
    'St. Louis Blues': 'NHL',
    'St. Louis Cardinals': 'MLB',
    'Tampa Bay Buccaneers': 'NFL',
    'Tampa Bay Lightning': 'NHL',
    'Tampa Bay Rays': 'MLB',
    'Tennessee Titans': 'NFL',
    'Texas Rangers': 'MLB',
    'Toronto Blue Jays': 'MLB',
    'Toronto FC': 'MLS',
    'Toronto Maple Leafs': 'NHL',
    'Toronto Raptors': 'NBA',
    'Utah Jazz': 'NBA',
    'Utah Mammoth': 'NHL',
    'Vancouver Canucks': 'NHL',
    'Vancouver Whitecaps FC': 'MLS',
    'Vegas Golden Knights': 'NHL',
    'Washington Capitals': 'NHL',
    'Washington Commanders': 'NFL',
    'Washington Mystics': 'WNBA',
    'Washington Nationals': 'MLB',
    'Washington Wizards': 'NBA',
    'Winnipeg Jets': 'NHL',
}

# Derive the team set from the league mapping (kept in sync)
MAJOR_LEAGUE_TEAMS = set(TEAM_LEAGUE.keys())

# Company → (sheet name, list of Company values in data)
COMPANY_SHEETS = {
    "Y&S":         ["YS Tickets", "YS-Seatgeek", "YS Tickets Spec", "YS-Seatgeek2"],
    "Grossman":    ["YSM Tickets"],
    "Sternbuch":   ["YSS Tickets"],
    "Pollak":      ["Pollak Tickets"],
    "Levine":      ["Yoni Levine"],
    "Levovitz":    ["Levovitz"],
    "GK":          ["GK LLC"],
    "Ticket Guy":  ["The Ticket Guy", "The Ticket Guy-Jas", "The Ticket Guy-Legacy", "The Ticket Guy VIP"],
    "Chase":       ["Jacks YS"],
    "Asher":       ["YSA", "YSA 2", "YSA 3"],
    "Katz":        ["YS Katz"],
    "TL":          ["YS TL"],
    "Waxler":      ["YSW"],
}


# ---------------------------------------------------------------------------
# Excel styling
# ---------------------------------------------------------------------------

def write_sheet(wb, name, dataframe):
    ws = wb.create_sheet(name)
    cols = list(dataframe.columns)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    fill_odd  = PatternFill("solid", start_color="FFFFFF")
    fill_even = PatternFill("solid", start_color="EEF2FF")

    for ri, row in enumerate(dataframe.itertuples(index=False), 2):
        row_fill = fill_even if ri % 2 == 0 else fill_odd
        for ci, val in enumerate(row, 1):
            col_name = cols[ci - 1]
            if col_name == "PO Created" and val is not None and not (isinstance(val, float) and np.isnan(val)):
                try:
                    val = pd.Timestamp(val).strftime("%m/%d/%Y")
                except Exception:
                    pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
            cell.fill = row_fill
            if col_name == "Total Cost":
                cell.number_format = "0.00"

    for ci, col in enumerate(cols, 1):
        max_len = len(str(col))
        for row in dataframe.itertuples(index=False):
            val = row[ci - 1]
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 55)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Core transformation
# ---------------------------------------------------------------------------

def apply_vendor_replacements(df):
    df["Vendor"] = df["Vendor"].replace("Box Office", "Default Vendor")
    df["Vendor"] = df["Vendor"].replace("Live Nation Flex", "Concert Seasons")
    for old, new in VENDOR_REPLACEMENTS:
        df["Vendor"] = df["Vendor"].str.replace(old, new, regex=False)
    return df


COLLEGE_KEYWORDS = ["college", "university", "football", "basketball", "hockey", "baseball"]


def league_for_team(team_performer):
    """Return the league label for a team, or '' if not a league/college team."""
    if not isinstance(team_performer, str):
        return ""
    # Exact match against the major league teams list
    if team_performer in TEAM_LEAGUE:
        return TEAM_LEAGUE[team_performer]
    # College detection via keywords
    low = team_performer.lower()
    if any(kw in low for kw in COLLEGE_KEYWORDS):
        return "College"
    return ""


def detect_season_ticket_keys(df_raw, min_event_dates=3):
    """
    Identify season-ticket combinations in raw source data.
    A combination of (Company, Team/Performer, Sec, Row, Seats, Email)
    that spans at least `min_event_dates` DISTINCT Event Dates is a season-ticket group.
    PO Created date and Total Cost are NOT part of the key, so groups can span
    multiple upload days and varying per-game prices.
    Rows with excluded vendors (resale marketplaces) are never counted.
    Returns a set of those key tuples.
    """
    needed = ["Company", "Team/Performer", "Sec", "Row", "Seats",
              "PO Email Account", "Event Date"]
    if any(c not in df_raw.columns for c in needed):
        return set()

    d = df_raw.copy()

    # Exclude resale-marketplace vendors from season-ticket detection
    if "Vendor" in d.columns:
        excluded_vendors = ["ticketmaster", "tickpick", "stubhub",
                            "ticket evolution", "gotickets"]
        v_low = d["Vendor"].astype(str).str.strip().str.lower()
        d = d[~v_low.isin(excluded_vendors)]

    # Normalize Event Date to a date for distinct-count
    d["_ev_key"] = fix_date(d["Event Date"]).dt.normalize().dt.date

    key_cols = ["Company", "Team/Performer", "Sec", "Row", "Seats",
                "PO Email Account"]
    # Build a string key, count distinct event dates per key
    grp = d.groupby(key_cols, dropna=False)["_ev_key"].nunique()
    season_keys = set(grp[grp >= min_event_dates].index)
    return season_keys


def season_league_map(df_raw, min_event_dates=3):
    """
    Returns a dict mapping the FULL season-ticket key
    (Company, Team/Performer, Sec, Row, Seats, Email) -> league label,
    for groups whose team has a league (major league or college).
    Only the specific seat group that meets the >=3 event-date threshold is labeled —
    other seat groups for the same team are NOT labeled (strict matching).
    """
    season_keys = detect_season_ticket_keys(df_raw, min_event_dates)
    result = {}
    for key in season_keys:
        team = key[1]  # key = (Company, Team/Performer, Sec, Row, Seats, Email)
        league = league_for_team(team)
        if league:
            result[key] = league
    return result


def build_all_query(df_raw):
    df = df_raw.copy()
    df["PO Created"] = fix_date(df["PO Created"]).dt.normalize().dt.date

    # ── Season-ticket detection (uses raw Event Date/Sec/Row/Seats before they're dropped) ──
    sl_map = season_league_map(df_raw, min_event_dates=3)
    # Tag each row with its season league using the FULL key
    # (Company, Team/Performer, Sec, Row, Seats, Email) so only the specific qualifying
    # seat group is labeled — not every row for that team.
    # A row only gets the tag if its own vendor is NOT an excluded resale marketplace.
    _excluded_vendors = {"ticketmaster", "tickpick", "stubhub", "ticket evolution", "gotickets"}
    _season_key_cols = ["Company", "Team/Performer", "Sec", "Row", "Seats", "PO Email Account"]
    if sl_map and all(c in df.columns for c in _season_key_cols):
        def _season_tag(r):
            if str(r["Vendor"]).strip().lower() in _excluded_vendors:
                return ""
            key = (r["Company"], r["Team/Performer"], r["Sec"], r["Row"],
                   r["Seats"], r["PO Email Account"])
            return sl_map.get(key, "")
        df["_SeasonLeague"] = df.apply(_season_tag, axis=1)
    else:
        df["_SeasonLeague"] = ""

    df["Ext PO #"] = df["Ext PO #"].fillna(" ").astype(str)
    df["PO Email Account"] = df["PO Email Account"].fillna(" ").astype(str)
    df = df.drop(columns=["Notes"], errors="ignore")
    df["Notes"] = (df["Team/Performer"].astype(str) + " / " +
                   df["PO Email Account"].astype(str) + " /  " +
                   df["Ext PO #"].astype(str))
    df["Notes (Short)"] = df["PO Email Account"].astype(str) + " / " + df["Ext PO #"].astype(str)

    drop_cols = ["PO #", "Opponent/Performer", "Event Date", "Sec", "Row", "Seats",
                 "Qty", "Cost", "Cancelled", "Created", "User", "Ext PO #", "PO Email Account"]
    df = df.drop(columns=[c for c in drop_cols if c in df.columns])

    df["Account"] = "Inventory Asset"
    df["Memo"] = df["Team/Performer"]
    df = df[["Company", "PO Created", "Account", "Vendor", "Team/Performer",
              "Memo", "Total Cost", "Notes", "Notes (Short)", "Delivery Type", "Venue", "Tags", "_SeasonLeague"]]

    # Broadway Groups → Broadway Seasons (before all other vendor logic)
    df["Vendor"] = df["Vendor"].replace("Broadway Groups", "Broadway Seasons")

    # YSA: Live Nation → Concert Seasons (then venue mapping handles the rest)
    ysa_companies = ["YSA", "YSA 2", "YSA 3"]
    df.loc[df["Company"].isin(ysa_companies) & (df["Vendor"] == "Live Nation"), "Vendor"] = "Concert Seasons"

    # Ticketmaster AM / Ballpark → team name (if major league team) or venue
    def resolve_tm_am(row):
        if row["Vendor"] not in ("Ticketmaster AM", "Ballpark"):
            return row["Vendor"]
        if row["Team/Performer"] in MAJOR_LEAGUE_TEAMS:
            return row["Team/Performer"]
        return row["Venue"]
    df["Vendor"] = df.apply(resolve_tm_am, axis=1)

    # Concert Extras at MSG / Beacon → Madison Square Garden (before Live Nation Extras rename)
    ce_msg_venues = ["Madison Square Garden", "Madison Square Garden Parking Lots",
                     "Beacon Theatre - New York"]
    ce_mask = (df["Vendor"] == "Concert Extras") & (df["Venue"].isin(ce_msg_venues))
    df.loc[ce_mask, "Vendor"] = "Madison Square Garden"

    df = apply_vendor_replacements(df)

    for old, new in [("Toyota Amphitheatre", "Live Nation Toyota Amp"),
                     ("Xfinity Center", "Live Nation Xfinity Center Boston"),
                     ("FPL Solar Ampitheater at Bayfront Park", "Live Nation FPL Solar Amp")]:
        df["Company"] = df["Company"].replace(old, new)
    df["Company"] = df["Company"].astype(str)
    df["Team/Performer"] = df["Team/Performer"].str.replace("Miami HEAT", "Miami Heat", regex=False)

    # Sports Extras → venue (but Radio City Music Hall → Madison Square Garden)
    sports_mask = df["Vendor"] == "Sports Extras"
    df.loc[sports_mask, "Vendor"] = df.loc[sports_mask].apply(
        lambda r: "Madison Square Garden" if r["Venue"] == "Radio City Music Hall" else r["Venue"], axis=1)

    # Ticket Guy broadway box office
    def ticket_guy_vendor(row):
        if row["Company"] not in ("The Ticket Guy", "The Ticket Guy-Jas", "The Ticket Guy-Legacy", "The Ticket Guy VIP"):
            return row["Vendor"]
        if row["Vendor"] != "Default Vendor":
            return row["Vendor"]
        if "New World Stages" in str(row["Venue"]):
            return "Box Office - New World Stages"
        return BROADWAY_VENUES.get(row["Venue"], row["Vendor"])
    df["Vendor"] = df.apply(ticket_guy_vendor, axis=1)

    # Concert Seasons
    df["Vendor"] = df.apply(
        lambda r: CONCERT_SEASONS_MAP.get(r["Venue"], r["Vendor"]) if r["Vendor"] == "Concert Seasons" else r["Vendor"], axis=1)

    # Broadway Seasons — collapse Team/Performer to "Various", rebuild Notes with "Various" as show name
    # Notes is already built as "ShowName / email / order#" — replace ShowName with "Various"
    is_bs = df["Vendor"] == "Broadway Seasons"
    df.loc[is_bs, "Team/Performer"] = "Various"
    df.loc[is_bs, "Memo"] = "Various"
    # Replace the show name portion of Notes (everything before the first " / ") with "Various"
    df.loc[is_bs, "Notes"] = df.loc[is_bs, "Notes"].str.replace(
        r"^.+?(?= / )", "Various", regex=True
    )
    df.loc[is_bs, "Notes (Short)"] = df.loc[is_bs, "Notes (Short)"]  # email/order# unchanged
    df["Vendor"] = df.apply(
        lambda r: BROADWAY_SEASONS_MAP.get(r["Venue"], r["Vendor"]) if r["Vendor"] == "Broadway Seasons" else r["Vendor"], axis=1)

    # MLB / Tickets.com
    df["MLB?"] = df["Team/Performer"].apply(lambda x: "Yes" if x in MLB_TEAMS else "No")
    df["Vendor"] = df.apply(
        lambda r: (r["Venue"] if r["MLB?"] == "No" else r["Team/Performer"]) if r["Vendor"] == "Tickets.com" else r["Vendor"], axis=1)
    df = df.drop(columns=["Venue", "MLB?"])

    # Proper case
    df["Vendor"] = df["Vendor"].str.title()
    df["Vendor"] = df["Vendor"].str.replace("Philadelphia 76Ers", "Philadelphia 76ers", regex=False)
    df["Vendor"] = df["Vendor"].str.replace("San Francisco 49Ers", "San Francisco 49ers", regex=False)

    # First groupby
    group_keys = ["Company", "PO Created", "Account", "Vendor", "Team/Performer",
                  "Memo", "Notes", "Notes (Short)", "_SeasonLeague"]
    df = df.groupby(group_keys, as_index=False, dropna=False)["Total Cost"].sum()
    df = df[df["Total Cost"] > 0]
    df["Vendor"] = df["Vendor"].astype(str)

    # Ticketmaster CAD
    df["VendorNew"] = df.apply(
        lambda r: ("Ticketmaster CAD" if str(r["Notes"]).endswith(("/TOR", "/VAN", "/QUE")) else "Ticketmaster")
        if r["Vendor"] == "Ticketmaster" else r["Vendor"], axis=1)
    df = df.rename(columns={"Notes": "Notes (Final)"})
    df = df.drop(columns=["Notes (Short)", "Vendor"])
    df = df.rename(columns={"VendorNew": "Vendor"})

    # Notes (Final) becomes Team/Performer (per M code)
    df = df[["Company", "PO Created", "Account", "Vendor", "Notes (Final)", "Total Cost", "_SeasonLeague"]]
    df = df.rename(columns={"Notes (Final)": "Team/Performer"})

    # Seasons tagging
    df["Seasons"] = df["Vendor"].apply(
        lambda v: "LN Extras" if v == "Live Nation Extras" else ("Live Nation" if "Live Nation" in str(v) else ""))
    df["Broadway_tag"] = df["Vendor"].apply(
        lambda v: "Broadway Extras" if v == "Broadway Extras" else ("Broadway" if "Broadway" in str(v) else ""))
    df["Seasons"] = df["Seasons"] + df["Broadway_tag"]
    df = df.drop(columns=["Broadway_tag"])

    # Apply season-ticket League label to rows not already tagged.
    # Skip rows whose FINAL vendor is an excluded resale marketplace — a raw vendor
    # like "Live Nation" can be renamed to "Ticketmaster", so the check must use the
    # final vendor name here, not the raw one captured earlier.
    _excluded_final_vendors = {"ticketmaster", "tickpick", "stubhub",
                               "ticket evolution", "gotickets"}
    def _apply_season(r):
        if r["Seasons"] == "" and r["_SeasonLeague"]:
            if str(r["Vendor"]).strip().lower() in _excluded_final_vendors:
                return ""
            return r["_SeasonLeague"]
        return r["Seasons"]
    df["Seasons"] = df.apply(_apply_season, axis=1)

    # Collapse LN seasons rows
    df["Team/Performer"] = df.apply(
        lambda r: "Various / Various" if r["Seasons"] in ("Live Nation", "LN Extras") else r["Team/Performer"], axis=1)
    df = df.rename(columns={"Team/Performer": "Memo2"})

    # Second groupby
    group_keys2 = ["Company", "PO Created", "Account", "Vendor", "Memo2", "Seasons"]
    df = df.groupby(group_keys2, as_index=False, dropna=False)["Total Cost"].sum()

    df["Team/Performer"] = df["Memo2"] + " (" + df["Company"] + ")"
    df["Memo"] = df["Team/Performer"]
    df["Bill No."] = [random.randint(10000000, 99999999) for _ in range(len(df))]

    return df[["Company", "Bill No.", "PO Created", "Account", "Vendor",
               "Memo2", "Team/Performer", "Memo", "Total Cost", "Seasons"]]


def build_summary_query(df_raw):
    s = df_raw.copy()
    s["PO Created"] = fix_date(s["PO Created"]).dt.normalize().dt.date
    s["Account"] = "Inventory Asset"
    # Broadway Groups → Broadway Seasons
    s["Vendor"] = s["Vendor"].replace("Broadway Groups", "Broadway Seasons")

    # YSA: Live Nation → Concert Seasons
    ysa_companies = ["YSA", "YSA 2", "YSA 3"]
    s.loc[s["Company"].isin(ysa_companies) & (s["Vendor"] == "Live Nation"), "Vendor"] = "Concert Seasons"

    # Ticketmaster AM / Ballpark → team name or venue
    def resolve_tm_am_s(row):
        if row["Vendor"] not in ("Ticketmaster AM", "Ballpark"):
            return row["Vendor"]
        if row["Team/Performer"] in MAJOR_LEAGUE_TEAMS:
            return row["Team/Performer"]
        return row["Venue"]
    s["Vendor"] = s.apply(resolve_tm_am_s, axis=1)

    # Concert Extras at MSG / Beacon → Madison Square Garden (before Live Nation Extras rename)
    ce_msg_venues_s = ["Madison Square Garden", "Madison Square Garden Parking Lots",
                       "Beacon Theatre - New York"]
    ce_mask_s = (s["Vendor"] == "Concert Extras") & (s["Venue"].isin(ce_msg_venues_s))
    s.loc[ce_mask_s, "Vendor"] = "Madison Square Garden"

    # Sports Extras at Radio City Music Hall → Madison Square Garden
    sports_mask_s = (s["Vendor"] == "Sports Extras") & (s["Venue"] == "Radio City Music Hall")
    s.loc[sports_mask_s, "Vendor"] = "Madison Square Garden"

    s = apply_vendor_replacements(s)
    s["Vendor"] = s["Vendor"].str.replace("FrontGate Tickets", "Front Gate Tickets", regex=False)
    s["MLB?"] = s["Team/Performer"].apply(lambda x: "Yes" if x in MLB_TEAMS else "No")
    s["Vendor"] = s.apply(
        lambda r: (r["Venue"] if r["MLB?"] == "No" else r["Team/Performer"]) if r["Vendor"] == "Tickets.com" else r["Vendor"], axis=1)

    grp = s.groupby(["Company", "PO Created", "Account", "Vendor", "Team/Performer"],
                    as_index=False, dropna=False).agg(Total_Cost=("Total Cost", "sum"), Qty=("Qty", "sum"))
    grp = grp.rename(columns={"Total_Cost": "Total Cost"})
    grp["Total Cost"] = grp["Total Cost"].round(2)
    grp = grp[["Company", "PO Created", "Account", "Vendor", "Team/Performer", "Qty", "Total Cost"]]
    return grp.sort_values(["Company", "PO Created", "Vendor", "Team/Performer"])


def filter_company(all_df, companies, rename_company=None, vendor_replace=None, strip_company_prefix=None):
    f = all_df[all_df["Company"].isin(companies)].copy()
    if rename_company:
        for old, new in rename_company.items():
            f["Company"] = f["Company"].str.replace(old, new, regex=False)
    if vendor_replace:
        for old, new in vendor_replace.items():
            f["Vendor"] = f["Vendor"].str.replace(old, new, regex=False)
    if strip_company_prefix:
        f["Company"] = f["Company"].str.replace(strip_company_prefix, "", regex=False)
    # Sort by PO Created > Vendor > Team/Performer (base name, ignoring email) > Total Cost
    sort_cols = [c for c in ["PO Created", "Vendor"] if c in f.columns]
    if "Team/Performer" in f.columns:
        # Extract base name (portion before first " / ") for sorting, ignoring email/order#
        f["_tp_sort"] = f["Team/Performer"].astype(str).str.split(" / ").str[0]
        sort_cols.append("_tp_sort")
    if "Total Cost" in f.columns:
        sort_cols.append("Total Cost")
    if sort_cols:
        f = f.sort_values(sort_cols, kind="mergesort").reset_index(drop=True)
    f = f.drop(columns=["_tp_sort"], errors="ignore")
    return f


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

# Maps the newer Purchase Details export schema onto the original column
# names the rest of the pipeline is built around.
NEW_FORMAT_COLUMN_MAP = {
    "CompanyName":        "Company",
    "POCreatedDate":      "PO Created",
    "Vendor":             "Vendor",
    "PrimaryEventName":   "Team/Performer",
    "SecondaryEventName": "Opponent/Performer",
    "VenueName":          "Venue",
    "Quantity":           "Qty",
    "CostPerTicket":      "Cost",
    "ExtPONumber":        "Ext PO #",
    "AccountEmail":       "PO Email Account",
    "DeliveryMethod":     "Delivery Type",
    "TextTagNames":       "Tags",
    "CreatedDate":        "Created",
    "CreatedBy":          "User",
}


def _is_new_format(df):
    """Detect the newer export schema by its signature columns."""
    cols = set(df.columns)
    return "CompanyName" in cols and "POCreatedDate" in cols


def _normalize_new_format(df):
    """Convert a newer-format DataFrame into the original schema in place."""
    df = df.rename(columns={k: v for k, v in NEW_FORMAT_COLUMN_MAP.items() if k in df.columns})

    # The new export's TotalCost column is unreliable (always 0), so derive the
    # line total from per-ticket cost x quantity.
    cost = pd.to_numeric(df.get("Cost"), errors="coerce").fillna(0.0)
    qty = pd.to_numeric(df.get("Qty"), errors="coerce").fillna(0.0)
    df["Total Cost"] = cost * qty

    # Cancellation is driven by IsPOCancelled OR IsCancelled in the new format.
    # Some rows have IsCancelled=true while IsPOCancelled=false — both must exclude the row.
    def _truthy(series):
        return series.astype(str).str.strip().str.lower().isin(["yes", "true", "1", "y"])

    cancel_flag = None
    if "IsPOCancelled" in df.columns:
        cancel_flag = _truthy(df["IsPOCancelled"])
    if "IsCancelled" in df.columns:
        ic = _truthy(df["IsCancelled"])
        cancel_flag = ic if cancel_flag is None else (cancel_flag | ic)
    if cancel_flag is not None:
        df["Cancelled"] = cancel_flag.map({True: "Yes", False: "No"})

    return df


def load_file(file_bytes, filename=""):
    """Load a single file (xlsx, xlsm, or csv) into a normalized DataFrame.

    Accepts either the original Purchase Details schema or the newer export
    schema, auto-detecting which one was uploaded and normalizing both to the
    column names the pipeline expects.
    """
    fname = filename.lower()
    if fname.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(file_bytes))
    else:
        xl = pd.ExcelFile(io.BytesIO(file_bytes))
        preferred = ["Source Data", "Input", "Template", "Sheet1", "Sheet"]
        sheet = next((s for s in preferred if s in xl.sheet_names), xl.sheet_names[0])
        df = xl.parse(sheet)

    if _is_new_format(df):
        df = _normalize_new_format(df)
    else:
        # Original-format column-name quirks (CSV vs Excel)
        col_map = {
            "Unnamed: 2": "Delivery Type",
            "Account Email": "PO Email Account",
            "Listing Created": "Created",
        }
        df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})

    # Ensure all expected columns exist (applies to both formats)
    for col in ["Delivery Type", "PO Email Account", "Ext PO #", "Notes", "Tags",
                "Cancelled", "Created", "User", "Opponent/Performer"]:
        if col not in df.columns:
            df[col] = None

    # Extract bill date from filename (pull date minus 1 day)
    m = DATE_FROM_FILENAME.search(filename)
    if m:
        pull_date = date_type(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        df["_bill_date"] = pull_date
    else:
        df["_bill_date"] = None

    return df


UUID_RE = re.compile(r'^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$')

def clean_ext_po(df):
    """Blank out UUID-format order numbers and order numbers for Concert Seasons / Ticketmaster AM rows."""
    if "Ext PO #" not in df.columns:
        return df

    def should_blank(v, vendor):
        if pd.isna(v):
            return False
        s = str(v).strip()
        # Always blank UUIDs and long numerics
        if UUID_RE.match(s):
            return True
        if s.isdigit() and len(s) >= 19:
            return True
        # Blank all Concert Seasons order numbers
        if vendor == "Concert Seasons":
            return True
        # Blank Ticketmaster AM order numbers 15+ chars
        if vendor == "Ticketmaster AM" and len(s) >= 15:
            return True
        return False

    new_vals = [
        None if should_blank(v, vend) else v
        for v, vend in zip(df["Ext PO #"], df["Vendor"])
    ]
    df["Ext PO #"] = pd.array(new_vals, dtype=object)
    return df


def process_file(file_bytes, filename=""):
    """Single file — wraps process_files for backwards compatibility."""
    return process_files([(file_bytes, filename)])



def convert_new_format(file_bytes, filename=""):
    """
    Convert a new-format TicketVault export to the old format layout.
    Returns bytes of an Excel file (.xlsx) with the reformatted data.
    """
    df = pd.read_excel(io.BytesIO(file_bytes))

    if not _is_new_format(df):
        raise ValueError("File does not appear to be in the new TicketVault format.")

    # Build Seats column from StartSeat-EndSeat
    def make_seats(row):
        try:
            start = int(row["StartSeat"])
            end = int(row["EndSeat"])
            if start == end:
                return str(start)
            return f"{start} - {end}"
        except:
            return ""

    df["Seats"] = df.apply(make_seats, axis=1)

    # Total Cost = CostPerTicket × Quantity
    df["Total Cost"] = pd.to_numeric(df["CostPerTicket"], errors="coerce").fillna(0) *                        pd.to_numeric(df["Quantity"], errors="coerce").fillna(0)

    # Compute combined cancellation (IsPOCancelled OR IsCancelled) before renaming,
    # matching Zone 2's logic so converted files carry the right Cancelled flag.
    def _truthy_conv(series):
        return series.astype(str).str.strip().str.lower().isin(["yes", "true", "1", "y"])
    conv_cancel = None
    if "IsPOCancelled" in df.columns:
        conv_cancel = _truthy_conv(df["IsPOCancelled"])
    if "IsCancelled" in df.columns:
        icc = _truthy_conv(df["IsCancelled"])
        conv_cancel = icc if conv_cancel is None else (conv_cancel | icc)
    if conv_cancel is not None:
        df["_CancelledCombined"] = conv_cancel.map({True: "Yes", False: "No"})

    # Map columns to old format order
    col_map = {
        "CompanyName":      "Company",
        "PurchaseOrderID":  "PO #",
        "POCreatedDate":    "PO Created",
        "Vendor":           "Vendor",
        "PrimaryEventName": "Team/Performer",
        "SecondaryEventName": "Opponent/Performer",
        "EventDateTime":    "Event Date",
        "VenueName":        "Venue",
        "Section":          "Sec",
        "Row":              "Row",
        "Quantity":         "Qty",
        "CostPerTicket":    "Cost",
        "ExtPONumber":      "Ext PO #",
        "AccountEmail":     "PO Email Account",
        "IsPOCancelled":    "Cancelled",
        "CreatedDate":      "Created",
        "UpdatedBy":        "User",
        "InternalNotes":    "Notes",
    }

    df = df.rename(columns=col_map)

    # Override Cancelled with the combined flag (covers IsCancelled-only cancellations)
    if "_CancelledCombined" in df.columns:
        df["Cancelled"] = df["_CancelledCombined"]
        df = df.drop(columns=["_CancelledCombined"])

    # Final column order (old format minus Delivery Type, Tags; includes Seats and Total Cost)
    final_cols = [
        "Company", "PO #", "PO Created", "Vendor", "Team/Performer",
        "Opponent/Performer", "Event Date", "Venue", "Sec", "Row", "Seats",
        "Qty", "Cost", "Total Cost", "Ext PO #", "PO Email Account",
        "Cancelled", "Created", "User", "Notes",
    ]
    # Keep only columns that exist
    final_cols = [c for c in final_cols if c in df.columns]
    df = df[final_cols]

    # Strip time portion from date columns — keep date only (MM/DD/YYYY)
    def date_only(v):
        if pd.isna(v):
            return v
        s = str(v).strip()
        if "T" in s:
            s = s.split("T")[0]
        try:
            return pd.Timestamp(s).strftime("%m/%d/%Y")
        except Exception:
            return v
    for date_col in ["PO Created", "Event Date", "Created"]:
        if date_col in df.columns:
            df[date_col] = df[date_col].apply(date_only)

    # Write to Excel with styling
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="4472C4")

    for ci, col in enumerate(final_cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    fill_odd  = PatternFill("solid", start_color="FFFFFF")
    fill_even = PatternFill("solid", start_color="EEF2FF")

    for ri, row in enumerate(df.itertuples(index=False), 2):
        row_fill = fill_even if ri % 2 == 0 else fill_odd
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val if val == val else None)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
            cell.fill = row_fill

    for ci, col in enumerate(final_cols, 1):
        max_len = max(len(col), df.iloc[:, ci-1].astype(str).str.len().max() if len(df) > 0 else 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 45)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(buf)
    return buf.getvalue()


def process_files(file_list):
    """
    Takes a list of (file_bytes, filename) tuples.
    Merges all into one DataFrame before processing.
    Returns a dict:
      {
        "date_range": "May 1 thru May 3 2026",
        "combined": <bytes of combined xlsx>,
        "companies": { "Y&S": <bytes>, ... },
        "stats": { ... },
        "all_companies": [...],
      }
    """
    frames = [load_file(fb, fn) for fb, fn in file_list]
    df_raw = pd.concat(frames, ignore_index=True)

    # Split out cancelled rows — exclude from processing but keep for Excluded tab
    def is_cancelled(val):
        if pd.isna(val): return False
        return str(val).strip().lower() in ("yes", "true", "1", "y")

    cancelled_mask = df_raw["Cancelled"].apply(is_cancelled)
    df_cancelled = df_raw[cancelled_mask].copy()
    df_raw = df_raw[~cancelled_mask].copy()

    # Override PO Created with bill date derived from filename (pull date - 1 day)
    has_bill_date = df_raw["_bill_date"].notna()
    if has_bill_date.any():
        df_raw["PO Created"] = df_raw.apply(
            lambda r: r["_bill_date"] if pd.notna(r["_bill_date"]) else r["PO Created"], axis=1
        )
    df_raw = df_raw.drop(columns=["_bill_date"], errors="ignore")

    # Clean Ext PO # — blank UUIDs and Concert Seasons / Ticketmaster AM order numbers
    df_raw = clean_ext_po(df_raw)

    # Determine date range
    dates = fix_date(df_raw["PO Created"]).dt.normalize().dropna().dt.date.unique()
    date_range_str = format_date_range([pd.Timestamp(d) for d in dates])

    # Build query outputs
    all_df = build_all_query(df_raw)
    summary_df = build_summary_query(df_raw)

    # Company filtered sheets
    company_dfs = {
        "Y&S":        filter_company(all_df, ["YS Tickets", "YS-Seatgeek", "YS Tickets Spec", "YS-Seatgeek2"]),
        "Grossman":   filter_company(all_df, ["YSM Tickets"]),
        "Sternbuch":  filter_company(all_df, ["YSS Tickets"]),
        "Pollak":     filter_company(all_df, ["Pollak Tickets"]),
        "Levine":     filter_company(all_df, ["Yoni Levine"]),
        "Levovitz":   filter_company(all_df, ["Levovitz"]),
        "Chase":      filter_company(all_df, ["Jacks YS"], rename_company={"Jacks YS": "Chase (Jacks)"}),
        "Asher":      filter_company(all_df, ["YSA", "YSA 2", "YSA 3"]),
        "Katz":       filter_company(all_df, ["YS Katz"]),
        "GK":         filter_company(all_df, ["GK LLC"], rename_company={"GK LLC": "YSKG"}),
        "TL":         filter_company(all_df, ["YS TL"]),
        "Waxler":     filter_company(all_df, ["YSW"], rename_company={"YSW": "YSW (Waxler)"}),
        "Ticket Guy": filter_company(all_df,
                          ["The Ticket Guy", "The Ticket Guy-Jas", "The Ticket Guy-Legacy", "The Ticket Guy VIP"],
                          vendor_replace={"Broadway Direct": "Box Office - Broadway Inbound"},
                          rename_company={"The Ticket Guy": "Ticket Guy", "The Ticket Guy-Jas": "Ticket Guy",
                                          "The Ticket Guy-Legacy": "Ticket Guy", "The Ticket Guy VIP": "Ticket Guy"}),
        "YourTickets":  filter_company(all_df, ["YourTickets"]),
    }

    # All company names (including empty) for UI display
    all_company_names = list(company_dfs.keys())

    # ── Build summary stats for UI display ───────────────────────────────────────
    stats = {}
    stats["Combined"] = {
        "rows": len(all_df),
        "total_cost": round(float(all_df["Total Cost"].sum()), 2),
    }
    for sheet_name, cdf in company_dfs.items():
        stats[sheet_name] = {
            "rows": len(cdf),
            "total_cost": round(float(cdf["Total Cost"].sum()), 2) if len(cdf) > 0 else 0.0,
        }

    return {
        "date_range": date_range_str,
        "all_companies": all_company_names,
        "stats": stats,
        # DataFrames for deferred filtered output building
        "_df_raw": df_raw,
        "_df_cancelled": df_cancelled,
        "_all_df": all_df,
        "_summary_df": summary_df,
        "_company_dfs": company_dfs,
    }


def build_filtered_outputs(df_raw, df_cancelled, all_df, summary_df, company_dfs, selected_companies, progress_cb=None):
    """Build combined workbook and per-company files for the selected companies only.
    progress_cb(done, total) is called as work completes."""
    selected_set = set(selected_companies)

    # Total steps = 1 (combined workbook) + number of non-empty selected companies
    selected_nonempty = [n for n, cdf in company_dfs.items() if n in selected_set and len(cdf) > 0]
    total_steps = 1 + len(selected_nonempty)
    done_steps = 0
    def _tick():
        nonlocal done_steps
        done_steps += 1
        if progress_cb:
            progress_cb(done_steps, total_steps)
    if progress_cb:
        progress_cb(0, total_steps)

    # Simpler: filter all_df by checking if rows belong to selected company sheets
    # Build a set of raw company values that map to selected sheet names
    # Map sheet name → raw Company values (true source names, before any renaming)
    raw_company_map = {
        "Y&S":        ["YS Tickets", "YS-Seatgeek", "YS Tickets Spec", "YS-Seatgeek2"],
        "Grossman":   ["YSM Tickets"],
        "Sternbuch":  ["YSS Tickets"],
        "Pollak":     ["Pollak Tickets"],
        "Levine":     ["Yoni Levine"],
        "Levovitz":   ["Levovitz"],
        "Chase":      ["Jacks YS"],
        "Asher":      ["YSA", "YSA 2", "YSA 3"],
        "Katz":       ["YS Katz"],
        "GK":         ["GK LLC"],
        "TL":         ["YS TL"],
        "Waxler":     ["YSW"],
        "Ticket Guy": ["The Ticket Guy", "The Ticket Guy-Jas", "The Ticket Guy-Legacy", "The Ticket Guy VIP"],
        "YourTickets":["YourTickets"],
    }

    # Build the set of raw Company values for the selected sheets.
    # Use raw_company_map (true source names) so renamed companies (e.g. Ticket Guy)
    # still match rows in the All/Summary tabs, which carry the raw names.
    selected_raw_companies = set()
    for sheet_name in selected_set:
        if sheet_name in raw_company_map:
            selected_raw_companies.update(raw_company_map[sheet_name])
        # Also include any raw values present in the company df (covers unmapped names)
        cdf = company_dfs.get(sheet_name)
        if cdf is not None and len(cdf) > 0:
            selected_raw_companies.update(cdf["Company"].unique())

    filtered_all = all_df[all_df["Company"].isin(selected_raw_companies)].copy()
    filtered_summary = summary_df[summary_df["Company"].isin(selected_raw_companies)].copy()

    # ── Build combined workbook ────────────────────────────────────────────────
    wb_combined = openpyxl.Workbook()
    wb_combined.remove(wb_combined.active)
    write_sheet(wb_combined, "Source Data", df_raw)
    if len(df_cancelled) > 0:
        write_sheet(wb_combined, "Canceled", df_cancelled)
    write_sheet(wb_combined, "All", filtered_all)
    write_sheet(wb_combined, "Summary", filtered_summary)
    for sheet_name, cdf in company_dfs.items():
        if sheet_name not in selected_set:
            continue
        write_sheet(wb_combined, sheet_name, cdf)
        if len(cdf) == 0:
            wb_combined[sheet_name].sheet_properties.tabColor = "FF0000"

    combined_buf = io.BytesIO()
    wb_combined.save(combined_buf)
    combined_bytes = combined_buf.getvalue()
    _tick()  # combined done

    # ── Build per-company workbooks (non-empty selected only) ──────────────────
    company_files = {}
    for sheet_name, cdf in company_dfs.items():
        if sheet_name not in selected_set or len(cdf) == 0:
            continue
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        write_sheet(wb, sheet_name, cdf)
        # Add filtered Input tab as second tab
        raw_companies = raw_company_map.get(sheet_name, [])
        company_input = df_raw[df_raw["Company"].isin(raw_companies)] if raw_companies else df_raw.iloc[0:0]
        if len(company_input) > 0:
            write_sheet(wb, "Source Data", company_input)
        buf = io.BytesIO()
        wb.save(buf)
        company_files[sheet_name] = buf.getvalue()
        _tick()  # this company done

    return combined_bytes, company_files
